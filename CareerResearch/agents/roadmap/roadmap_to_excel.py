from openpyxl import Workbook
from io import BytesIO

def roadmap_to_excel(roadmap_text: str) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "60-Day Career Roadmap"

    headers = [
        "Week",
        "Task",
        "Skill Focus",
        "Estimated Effort (hrs)",
        "Outcome",
        "Status",
        "Notes"
    ]
    ws.append(headers)

    current_week = ""

    for line in roadmap_text.splitlines():
        line = line.strip()

        if line.lower().startswith("week"):
            current_week = line.replace(":", "")
        elif line.startswith("- Task:"):
            task = line.replace("- Task:", "").strip()
            ws.append([current_week, task, "", "", "", "Not Started", ""])
        elif line.startswith("Skill Focus:"):
            ws.cell(row=ws.max_row, column=3).value = line.replace("Skill Focus:", "").strip()
        elif line.startswith("Estimated Effort"):
            ws.cell(row=ws.max_row, column=4).value = line.split(":")[-1].strip()
        elif line.startswith("Outcome:"):
            ws.cell(row=ws.max_row, column=5).value = line.replace("Outcome:", "").strip()

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
